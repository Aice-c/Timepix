"""Excel workbook export for thesis analysis tables."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd


def sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]\*:/\\?]", "_", name)
    return cleaned[:31] or "Sheet"


def write_analysis_workbook(
    tables: list[tuple[str, str, str, pd.DataFrame]],
    out_path: str | Path,
    *,
    title: str,
) -> Path:
    """Write all analysis tables into one xlsx workbook with readable sheets.

    Each tuple is `(sheet_name, table_title, table_note, dataframe)`.
    """

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        import openpyxl  # noqa: F401
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:  # pragma: no cover - dependency guidance
        raise ImportError("openpyxl is required to export multi-sheet analysis workbooks") from exc

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        cover = pd.DataFrame(
            [
                {"项目": "工作簿标题", "内容": title},
                {"项目": "说明", "内容": "本工作簿汇总论文数据分析脚本生成的统计表。图形输出见对应 figures 目录。"},
                {"项目": "字体约定", "内容": "表题、表注使用中文；专业字段保留英文。论文排版时按三线表重新整理。"},
            ]
        )
        cover.to_excel(writer, sheet_name="说明", index=False, startrow=2)
        for raw_name, table_title, table_note, df in tables:
            df = df.copy()
            name = sheet_name(raw_name)
            df.to_excel(writer, sheet_name=name, index=False, startrow=3)
            ws = writer.book[name]
            ws["A1"] = table_title
            ws["A2"] = table_note

        workbook = writer.book
        thin_gray = Side(style="thin", color="D9D9D9")
        header_fill = PatternFill("solid", fgColor="F2F2F2")
        title_fill = PatternFill("solid", fgColor="D9EAF7")
        for ws in workbook.worksheets:
            ws.freeze_panes = "A4"
            max_col = ws.max_column
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, max_col))
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, max_col))
            ws["A1"].font = Font(name="SimSun", bold=True, size=12)
            ws["A1"].fill = title_fill
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A2"].font = Font(name="SimSun", italic=True, size=10, color="666666")
            ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
            header_row = 4
            for cell in ws[header_row]:
                cell.font = Font(name="SimSun", bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(top=thin_gray, bottom=thin_gray)
            for row in ws.iter_rows(min_row=5):
                for cell in row:
                    cell.font = Font(name="Times New Roman", size=10)
                    cell.alignment = Alignment(vertical="center")
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                lengths = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
                width = min(max(max(lengths, default=8) + 2, 10), 35)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.row_dimensions[1].height = 24
            ws.row_dimensions[2].height = 36
    return out_path
